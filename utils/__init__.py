from google.genai import types
from google import genai
from curl_cffi import requests as cloaked_requests
import os
import pathlib
from random import uniform
import re
from time import sleep
import time
from curl_cffi import ProxySpec, requests
import random
from pathlib import Path
from os.path import join
import openpyxl.styles
import ua_generator
from ua_generator.options import Options as OptionsUA
from ua_generator.data.version import VersionRange
from datetime import datetime
import openpyxl
import json
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import io
import os


def isin_from_gemini(input: str):
    client = genai.Client(
        api_key=os.environ.get("GEMINI_API_KEY")
    )

    model = "gemini-3.1-flash-lite"
    contents = [
        types.Content(
            role="user",
            parts=[
                types.Part.from_text(
                    text=input)
            ],
        ),
    ]
    tools = [
        types.Tool(url_context=types.UrlContext()),
    ]
    generate_content_config = types.GenerateContentConfig(
        thinking_config=types.ThinkingConfig(
            thinking_level="MINIMAL",
        ),
        tools=tools,
        system_instruction=[
            types.Part.from_text(text="""* Extract only first ISIN found withing input url.
* Stop further processing once an ISIN was found.
* Return None if nothing was found."""),
        ],
    )

    out = []
    for chunk in client.models.generate_content_stream(
        model=model,
        contents=contents,
        config=generate_content_config,
    ):
        if text := chunk.text:

            out.append(text)
    return "".join(out)


def get_random_proxy_port_str():
    return f"socks5://c23aa2273d4cf55a8726__cr.gb:be209b0843f58c7e@gw.dataimpulse.com:10{random.randint(0, 999):03}"


def get_proxy_endpoint() -> dict[str, str]:
    # Base port configuration (e.g., mapping workers to ports 10001, 10002, etc.)

    attempt = 1
    while True:
        # 1. Build the clean standard format string for CloakBrowser
        # browser_proxy_str = f"socks5://{username}:{password}@{endpoint_domain}:{target_port}"
        browser_proxy_str = get_random_proxy_port_str()

        # 2. Convert to 'socks5h://' to force curl_cffi to use remote DNS lookups
        curl_proxy_str = browser_proxy_str
        socks_proxies = ProxySpec(
            {"http": curl_proxy_str, "https": curl_proxy_str})

        print(
            f"🔄 [Proxy Check] Testing port {browser_proxy_str} for... (Attempt {attempt})")

        try:
            # Quick network request with a strict 5-second timeout
            response = cloaked_requests.get(
                "https://api.ipify.org",
                proxies=socks_proxies,
                impersonate="chrome",
                timeout=5
            )

            if response.status_code == 200 and response.text.strip():
                print(
                    f"✅ [Proxy Live] passed check. Exit IP: {response.text.strip()}")
                # Returns the pristine socks5:// format for CloakBrowser
                return dict(ip=response.text.strip(), proxy=browser_proxy_str)

        except Exception as e:
            print(f"⚠️ [Port Down] Port failed handshake: {e}")

        # 3. Apply a small backoff pause before cycling the loop
        sleep_time = min(attempt * 3, 10)
        print(f"⏳ Sleeping {sleep_time}s before re-checking port...")
        time.sleep(sleep_time)
        attempt += 1


def get_xlsx_data(filename, sheet_name) -> list[dict]:
    wb = openpyxl.load_workbook(filename)
    ws = wb[sheet_name]
    data = []
    row_start = 2
    for row in range(row_start, ws.max_row + 1):
        if ws.cell(row, 2).value:
            continue
        c1 = ws.cell(row, 1).value
        c3 = ws.cell(row, 3).value
        f = dict(name=c1, url=c3)
        data.append(f)
    wb.close()
    return data


def create_spreadsheet(filename, sheet_names, column_names, col_width=25):
    wb = openpyxl.Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    header_font = Font(name='Arial', size=12, bold=True, color="000000")
    header_fill = PatternFill(
        start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold background

    for name in sheet_names:
        ws = wb.create_sheet(title=name)
        ws.append(column_names)

        # Apply style and set column width
        for i, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.fill = header_fill

            # Use the column letter (A, B, C...) to set width
            column_letter = get_column_letter(i)
            ws.column_dimensions[column_letter].width = col_width

    wb.save(filename)


def isin_from_text(text: str) -> str | None:

    # 1. Relaxed regex to find ISINs even if they have a random space inside
    isin_extract_rx = re.compile(
        r"[A-Z]{2}(?:[?\s]*[A-Z0-9]){9}[?\s]*[0-9]")
    # 2. Strict regex to validate after cleaning
    isin_strict_rx = re.compile(r"^[A-Z]{2}[A-Z0-9]{9}[0-9]$")
    matches = isin_extract_rx.findall(text)

    for match in matches:
        # Clean the extracted string by removing all spaces
        cleaned_isin = match.replace(" ", "")

        # Strictly validate
        if isin_strict_rx.match(cleaned_isin):
            return cleaned_isin
    return None


def get_random_user_agent(platform: list[str] = ["windows"]) -> dict:
    options = OptionsUA()
    options.version_ranges = {
        "chrome": VersionRange(140, 144),  # Choose version between 125 and 129
    }
    ua = ua_generator.generate(
        browser="chrome", platform=platform, options=options)
    ua.headers.accept_ch(
        "Sec-CH-UA-Platform-Version, Sec-CH-UA-Full-Version-List")
    # return ua.headers.get()
    headers = ua.headers.get()
    return {k.title(): v for k, v in headers.items()}


def delay(min: float, max: float):
    sleep(uniform(min, max))


def clean_spreadsheet(filename: str) -> None:
    wb = openpyxl.load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
            for cell in row:
                cell.value = None
                cell.hyperlink = None
    wb.save(filename)
    wb.close()


def get_current_quarter(date_obj) -> str:
    return f"Q{(date_obj.month - 1) // 3 + 1}"


def email_title() -> str:
    now = datetime.now()
    return f"[Automation] {get_current_quarter(now)} report for Halifax - Chelsea Financial - iWeb - Quilter - Standard Life - Willis Owen"


def get_xlsx_filepath(filename: str) -> str:
    project_root = Path(__file__).resolve().parent.parent
    return join(project_root, "spreadsheet", filename)


def write_json(filename: str, data: list[dict]) -> None:
    with open(filename, "w+") as f:
        json.dump(data, f, indent=4)


def read_json(filename: str):
    project_root = Path(__file__).resolve().parent.parent
    json_dir = os.path.join(project_root, "json")
    if pathlib.Path.is_dir(Path(json_dir)):
        filename = os.path.join(json_dir, filename)

        if pathlib.Path.exists(Path(filename)):
            with open(filename, "r") as f:
                return json.load(f)
        print("Filename doesn't exist inside json folder: ", filename)
    print("json folder doesn't exist: ", filename)


def get_fund_type_total(fund_type: str) -> list[int]:
    project_root = Path(__file__).resolve().parent.parent
    json_path = join(project_root, "json", "total.json")
    data = read_json(json_path)
    if data:
        for d in data:
            if d["name"] == fund_type:
                return [i for i in range(1, d["total"]+1)]
    return []


def fetch_with_backoff(
    url, headers=get_random_user_agent(), cookies=None, max_retries=5, base_delay=2, proxy=""
):
    for attempt in range(max_retries):
        try:
            # Using curl_cffi to mimic a real browser (e.g., Chrome)
            proxy_spec = ProxySpec(
                {"http": proxy, "https": proxy}
            )
            response = requests.get(
                url, headers=headers, cookies=cookies, impersonate="chrome", timeout=10, proxies=proxy_spec
            )

            # If successful, return the response
            if response.status_code == 200:
                return response

            # If we hit rate limits or server errors, we should retry
            if response.status_code in [429, 500, 502, 503, 504]:
                print(
                    f"Attempt {attempt + 1} failed with status {response.status_code}. Retrying..."
                )
            else:
                # For 404 or 403, retrying usually won't help
                print(
                    f"Permanent error {response.status_code}. Skipping retries.")
                return response

        except Exception as e:
            print(f"Attempt {attempt + 1} raised an exception: {e}")

        jitter = random.uniform(0.5, 1.5)
        sleep_time = (base_delay * (2**attempt)) * jitter

        print(f"Sleeping for {sleep_time:.2f} seconds...")
        time.sleep(sleep_time)

    print("Max retries reached. Mission failed.")
    return None


def save_xlsx(

    xlsx_out: str,
    funds: list[dict],
    cols: list[str],
    sheet: str,
    start: int = 2,
):
    wb = openpyxl.load_workbook(xlsx_out)
    ws = wb[sheet]
    for fund in funds:
        for idx, val in enumerate(cols):
            col = idx + 1
            row = fund.get("index")
            if row:
                start = int(row)
            if val == "url":
                cell = ws.cell(start, col, fund.get(val))
                cell.style = "Hyperlink"
                cell.hyperlink = fund.get(val)
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
                continue
            ws.cell(start, col, fund.get(val))
        start += 1
    wb.save(xlsx_out)
    wb.close()


def extract_isin_from_pdf_bytes(pdf_bytes: bytes) -> str | None:
    """Parses raw PDF bytes in-memory and executes loose-and-strict regex matching to extract the ISIN."""
    try:
        pdf_file = io.BytesIO(pdf_bytes)
        reader = PdfReader(pdf_file)

        # Compile your regex engines (matching your exact kiid.py implementation)
        isin_extract_rx = re.compile(
            r"[A-Z]{2}(?:[?\s]*[A-Z0-9]){9}[?\\s]*[0-9]")
        isin_strict_rx = re.compile(r"^[A-Z]{2}[A-Z0-9]{9}[0-9]$")

        # Iterate through pages to pull out text buffers
        for page in reader.pages:
            text = page.extract_text() or ""
            matches = isin_extract_rx.findall(text)

            for match in matches:
                cleaned_isin = match.replace(" ", "")
                if isin_strict_rx.match(cleaned_isin):
                    return cleaned_isin
    except Exception as e:
        print(f"Error parsing PDF structures: {e}")
    return None
